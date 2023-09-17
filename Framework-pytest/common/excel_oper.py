"""
报错解析：
  ‘ File "D:\python_3.9.8\lib\zipfile.py", line 1239, in __init__
     self.fp = io.open(file, filemode)
    PermissionError: [Errno 13] Permission denied: 'E:\\MessageNotification\\Api\\date.xlsx'
    =====这个报错是因为Excel为打开状态，关闭即可

"""
import openpyxl


class Excel_Operation:
    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):
        """数据读取的方法"""
        """
        filename:读取的Excel文件路径
        sheetname:需要读取的sheet页名称
        .value：通过该属性获取格子内容
        """
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        res = list(sh.rows)
        # 获取第一行数据
        title = [i.value for i in res[0]]
        cases = []
        # 遍历除第一行之外的所有行数据
        for item in res[1:]:
            data = [i.value for i in item]
            dic = dict(zip(title, data))
            cases.append(dic)
        return cases

    def write_data(self, row, column, value):
        """数据写入的方法"""
        '''
        写入数据的位置
        row：行
        column：列
        value：值/内容
        '''
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        # 写入数据
        sh.cell(row=row, column=column, value=value)
        workbook.save(self.filename)


if __name__ == '__main__':
    # 调用方法，向Excel中写入数据
    ex = Excel_Operation(r'E:\MessageNotification\Api\date.xlsx', 'Sheet2')
    res = ex.write_data(row=7,column=10,value='写入数据内容')

    # 调用方法，读取Excel数据
    ex = Excel_Operation(r'E:\MessageNotification\Api\date.xlsx', 'Sheet2')
    res1 = ex.read_data()
    print(res1)

